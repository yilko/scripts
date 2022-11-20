import configparser, hashlib, os;
from datetime import datetime;
from openpyxl import Workbook, load_workbook;
from openpyxl.styles import Font, colors, Alignment;
from config_log import ConfigLog;

'''实现文件信息保存到excel及md5保存到txt文件'''
log = ConfigLog().config_log();


class SaveFiles:

    def __init__(self):
        # 初始化对象
        config = configparser.ConfigParser();
        # 读取配置文件
        config.read("file_config.ini", encoding="utf-8-sig");
        self.create_folder = config.get("create_file", "folder_name");
        self.get_md5 = config.getboolean("md5_file", "get_md5");
        self.save_folder = config.get("md5_file", "folder_name");

    # 获取文件信息返回列表
    def __get_file_md5(self):
        folder = self.create_folder;
        # 获取文件夹下所有文件的内容（列表）
        file_ls = os.listdir(folder);
        ls = [];
        for file_name in file_ls:
            # 文件的具体路径
            file_path = f"{self.create_folder}/{file_name}"
            # 获取大小
            file_size = str(os.path.getsize(file_path));
            # 获取md5
            with open(file_path, "rb") as f:
                content = f.read();
            file_md5 = hashlib.md5(content).hexdigest();
            file_info_ls = [file_name, file_size, file_md5, file_path];
            ls.append(file_info_ls);
        log.debug(f"当前目录下有文件{len(ls)}个,分别是{ls}");
        return ls;

    # 保存到excel
    def __save_excel(self, file_ls: list):
        # 判断save文件夹是否包含xlsx
        for file in os.listdir(self.save_folder):
            if file == "文件信息.xlsx":
                wb = load_workbook(f"{self.save_folder}/文件信息.xlsx");
                break;
        # 不包含该文件，创一个新的
        else:
            wb = Workbook();
        log.info("当前目录下文件信息正在保存到excel...")
        # 当前日期作为工作表名字
        str_curr_date = str(datetime.now().date());
        # 出现工作表同名情况+时间戳
        for sheet_name in wb.sheetnames:
            if sheet_name == str_curr_date:
                str_curr_date = str_curr_date + "+" + str(datetime.now().timestamp());
                break;
        log.info(f"当前工作表名称为===={str_curr_date}");
        # 0代表插入到最前面的位置
        ws = wb.create_sheet(str_curr_date, 0);
        title_ls = ["文件名", "文件大小(b)", "文件md5", "文件路径"];
        for i in range(1, len(title_ls) + 1):
            ws.cell(1, i).value = title_ls[i - 1];
        # 设置列宽
        col_letters = ["A", "B", "C", "D"];
        width_nums = [40, 13, 35, 60];
        for i in range(len(col_letters)):
            letter = col_letters[i]
            ws.column_dimensions[letter].width = width_nums[i];
        # 处理第一行的标题样式
        title_font = Font(name="微软雅黑", size=12, bold=True);
        title_align = Alignment(horizontal="center", vertical="center");
        for col in range(1, ws.max_column + 1):
            ws.cell(1, col).font = title_font;
            ws.cell(1, col).alignment = title_align;
        # 把获取到的信息一一写到对应的单元格中
        for row in range(2, len(file_ls) + 2):
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).value = file_ls[row - 2][col - 1];
                ws.cell(row, col).alignment = title_align;
        save_path = f"{self.save_folder}/文件信息.xlsx";
        wb.save(save_path);
        log.info(f"保存excel成功！路径在===={save_path}");

    # 保存到txt
    def __save_txt(self, file_ls: list):
        save_path = f"{self.save_folder}/md5.txt";
        with open(save_path, "w", encoding="utf-8-sig") as f:
            for file in file_ls:
                f.write(f"{file[2]}\n");
        log.info(f"保存txt成功！路径在===={save_path}");

    # 获取md5并保存excel和txt
    def save_md5(self):
        if self.get_md5:
            if os.path.exists(self.create_folder):
                file_ls = self.__get_file_md5();
                if file_ls:
                    folder = self.save_folder;
                    if not os.path.exists(folder):
                        os.mkdir(folder);
                    self.__save_excel(file_ls);
                    self.__save_txt(file_ls);
                else:
                    log.info(f"不需要保存数据到文件中，列表为空===={file_ls}");
            else:
                log.info(f"当前路径下没有创建{self.create_folder}文件夹，没有文件可获取md5");
        else:
            log.info("不需要获取文件md5信息");
