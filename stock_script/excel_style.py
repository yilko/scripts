from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font


class ExcelStyle:
    def __init__(self, excel_path):
        self.path = excel_path
        self.wb: Workbook = load_workbook(excel_path)
        self.sheet_ls = self.wb.sheetnames

    # 单个工作表自适应列宽
    @staticmethod
    def adjust_width(ws: Worksheet):
        count = 0
        for col in ws.columns:
            count += 1
            col_ls = []
            for cell in col:
                value = cell.value
                # 存在精度问题，防止单元格长度过长
                if isinstance(value, (int, float)):
                    value = round(value, 2)
                col_ls.append(len(str(value)))
            max_len = max(col_ls)
            ws.column_dimensions[get_column_letter(count)].width = max_len * 1.3 + 5

    # 调整居中对齐方式和字体样式
    @staticmethod
    def adjust_align_font(ws: Worksheet, horizontal: str, vertical: str, font_name: str, font_size: int):
        for row in ws.rows:
            for cell in row:
                cell.alignment = Alignment(horizontal, vertical)
                cell.font = Font(font_name, font_size)

    def adjust_style(self):
        for sheet in self.sheet_ls:
            ws = self.wb[sheet]
            self.adjust_width(ws)
            # 冻结第一行,A2=A1-Z1被冻结，A2以下正常滚动
            ws.freeze_panes = 'A2'
            self.adjust_align_font(ws, 'center', 'center', '微软雅黑', 11)
        self.wb.save(self.path)


if __name__ == '__main__':
    es = ExcelStyle('20241101.xlsx')
    es.adjust_style()
